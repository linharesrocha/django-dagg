from django.shortcuts import render, redirect
from django.http import HttpResponse
from scripts import planilha_campanha, produtos_sem_venda, comparativo_vendas_netshoes, relatorio_envio_magalu, todas_vinculacoes_aton_marketplace, todas_as_vendas_aton, relatorio_envio_full_v2, pedidos_do_dia, cria_planilha_armazem_valor_custo_total
from datetime import datetime, date
from django.contrib.messages import constants
from django.contrib import messages
from io import BytesIO
from openpyxl.utils import get_column_letter
from relatorios.scripts import produtos_mlb_stats_pedro
import os
from scripts.connect_to_database import get_connection
import pyodbc
from openpyxl.styles import PatternFill
import pandas as pd
import warnings

def index(request):
    return render(request, 'index-relatorios.html')

def gerar_planilha_campanha(request):
    
    # Obtem valor marketplace
    marketplace = request.POST['marketplace']
    
    if marketplace == 'Tudo':
        marketplace = None
    
    # Criando o nome da planilha com data e hroa
    today = date.today()
    dia_atual = str(today.strftime("%Y-%m-%d"))
    agora_hora = datetime.now()
    hora_atual = str(agora_hora.strftime("%H:%M:%S")).replace(':','-')
    nome_planilha = f'planilha-de-campanha-{dia_atual}-{hora_atual}.xlsx'
    
    # Chame a função que gera o arquvio Excel
    arquivo_excel = planilha_campanha.main(marketplace=marketplace)
    
    # Crie uma resposta HTTP para retornar o arquivo ao usuário
    response = HttpResponse(arquivo_excel, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{nome_planilha}"'
    return response

def gerar_planilha_produtos_sem_venda(request):
    
    # Criando o nome da planilha
    today = date.today()
    last_month = date(today.year, today.month-1, 1)
    month_name = last_month.strftime('%B')
    nome_planilha = f'produtos-sem-venda-{month_name}.xlsx'
    
    # Chame a função que gera o arquvio Excel
    arquivo_excel = produtos_sem_venda.main()
    
    response = HttpResponse(arquivo_excel, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename={nome_planilha}'
    return response

def gerar_planilha_comparativo_vendas_netshoes(request):
    
    hoje = date.today()
    hoje_formatado = hoje.strftime("%Y-%m-%d")
    
    if request.method == 'POST':
        data_inicial_principal = request.POST['data_inicial_principal']
        data_final_principal = request.POST['data_final_principal']
        data_inicial_comparativo = request.POST['data_inicial_comparativo']
        data_final_comparativo = request.POST['data_final_comparativo']
        
        # Verifica se todas estão com data
        if data_inicial_principal == '' or data_final_principal == '' or data_inicial_comparativo == '' or data_inicial_comparativo == '':
            messages.add_message(request, constants.ERROR, 'Preencha todas as datas!')
            return redirect('index-relatorios')
        
        # Valida se a data é até hoje
        if data_final_principal > hoje_formatado:
            messages.add_message(request, constants.ERROR, 'Data principal final maior que hoje!')
            return redirect('index-relatorios')
        
        # Valida data principal 
        if data_inicial_principal > data_final_principal:
            messages.add_message(request, constants.ERROR, 'Data principal inicial maior que a data principal final!')
            return redirect('index-relatorios')
        
        # Valida data comparativa
        if data_inicial_comparativo > data_final_comparativo:
            messages.add_message(request, constants.ERROR, 'Data comparativa inicial maior que a data comparativa final!')
            return redirect('index-relatorios')
        
        # Nome planilha
        nome_planilha = f'relatorio_comparativo_netshoes.xlsx'
        
        # Chama função de relatório
        arquivo_excel = comparativo_vendas_netshoes.main(data_inicial_principal=data_inicial_principal,
                                         data_final_principal=data_final_principal,
                                         data_inicial_comparativo=data_inicial_comparativo,
                                         data_final_comparativo=data_final_comparativo)
        
        response = HttpResponse(arquivo_excel, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename={nome_planilha}'
        return response
    
def gerar_planilha_todas_vinculacoes_aton_marketplace(request):
    nome_planilha = f'todas_vinculacoes_aton_marketplace.xlsx'
    marketplace = request.POST['marketplace']

    if marketplace == 'Tudo':
        marketplace = None
    
    arquivo_excel = todas_vinculacoes_aton_marketplace.main(marketplace=marketplace)
    
    # Crie uma resposta HTTP para retornar o arquivo ao usuário
    response = HttpResponse(arquivo_excel, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{nome_planilha}"'
    return response

def gerar_planilha_todas_as_vendas_aton(request):
    arquivo_excel = todas_as_vendas_aton.main()
    
     # Crie uma resposta HTTP para retornar o arquivo ao usuário
    response = HttpResponse(arquivo_excel, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="todas_as_vendas_aton.xlsx"'
    return response

def gerar_planilha_envio_full(request):
    file = request.FILES['file']
    
    output = relatorio_envio_full_v2.main(file)
    
    # Retorne a resposta HTTP com o arquivo Excel como anexo
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=analise_envio_full.xlsx'
    return response

def gerar_planilha_magalu_full(request):
    file = request.FILES['file']
    empresa_personalizada = request.POST['empresa_personalizada']

    output = relatorio_envio_magalu.main(file, empresa_personalizada)

    # Retorne a resposta HTTP com o arquivo Excel como anexo
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=analise_envio_full.xlsx'
    return response

def gerar_planilha_peso_quant(request):
    file = request.FILES['file']
    
    df_peso_quant = pd.read_excel(file)
    
    try:
        df_peso_quant['COD_INTERNO'] = df_peso_quant['COD_INTERNO'].str.strip()
    except:
        messages.add_message(request, constants.ERROR, 'Erro ao ler planilha! Verifique se a primeira coluna se chama "COD_INTERNO" e a segunda coluna se chama "QUANT" ou "Quant."')
        return redirect('index-relatorios')
    
    # Verifica se a primeira coluna se chama COD_INTERNO
    if str(df_peso_quant.columns[0]).upper() != 'COD_INTERNO':
        messages.add_message(request, constants.ERROR, 'A primeira coluna deve se chamar "COD_INTERNO"')
        return redirect('index-relatorios')
    
    # Verifica se a segunda coluna se chama QUANT
    if df_peso_quant.columns[1].upper() != 'QUANT' and str(df_peso_quant.columns[1]) != 'Quant.':
        messages.add_message(request, constants.ERROR, 'A segunda coluna deve se chamar "QUANT" ou "Quant."')
        return redirect('index-relatorios')
    
    df_peso_quant.rename(columns={'COD_INTERNO': 'COD_INTERNO', 'quant': 'QUANT', 'Quant.': 'QUANT'}, inplace=True)
    
    try:
        # Puxa Peso
        lista_cod_interno = df_peso_quant['COD_INTERNO'].tolist()
        cod_interno_com_aspas = ["'" + str(cod) + "'" for cod in lista_cod_interno]
        cod_interno_str = ', '.join(cod_interno_com_aspas)
        warnings.filterwarnings('ignore')
        connection = get_connection()
        conexao = pyodbc.connect(connection)
        
        comando = f'''
        SELECT COD_INTERNO, PESO
        FROM MATERIAIS
        WHERE COD_INTERNO IN ({cod_interno_str})
        '''
        
        df_resultado_peso = pd.read_sql(comando, conexao)
        df_resultado_peso['COD_INTERNO'] = df_resultado_peso['COD_INTERNO'].str.strip()
        
        df_peso_quant = pd.merge(df_peso_quant, df_resultado_peso, on='COD_INTERNO', how='left')
        
        # Multiplica coluna QUANT por PESO
        df_peso_quant['PESO_MULT'] = df_peso_quant['QUANT'] * df_peso_quant['PESO']
        
        # Soma todos os pesos
        peso_total = df_peso_quant['PESO_MULT'].sum()
        
        # Cria coluna para colocar peso_total mas apenas na primeira celula
        df_peso_quant['PESO_TOTAL'] = peso_total
        
        excel_bytes = BytesIO()
        df_peso_quant.to_excel(excel_bytes, index=False)
        excel_bytes.seek(0)
        bytes_data = excel_bytes.getvalue()
        
        response = HttpResponse(bytes_data, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="peso_quant.xlsx"'
        return response
    except:
        messages.add_message(request, constants.ERROR, 'Erro ao gerar planilha!')
        return redirect('index-relatorios')
        
    

def gerar_planilha_pedidos_dia(request):
    output = pedidos_do_dia.main()
    
    today = date.today()
    dia_atual = str(today.strftime("%d-%m-%Y"))
    nome_planilha = f'pedidos_do_dia_{dia_atual}'
    
    # Retorne a resposta HTTP com o arquivo Excel como anexo
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={nome_planilha}.xlsx'
    return response

def gerar_mlbs_stats(request):
    arquivo_excel = produtos_mlb_stats_pedro.main()
    
    # Crie uma resposta HTTP para retornar o arquivo ao usuário
    response = HttpResponse(arquivo_excel, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="mlbs_stats.xlsx"'
    return response

def margem_netshoes_personalizada(request):
    empresa_personalizada = request.POST['empresa_personalizada']
    data_inicio = request.POST['data_inicial']
    data_fim = request.POST['data_final']
    personalizado = True

    print(data_inicio, data_fim)

    from scripts.confere_margem_netshoes import main

    main(data_inicio, data_fim, empresa_personalizada, personalizado)

    return redirect('index-relatorios')

def margem_decathlon_personalizada(request):
    data_inicio = request.POST['data_inicial']
    data_fim = request.POST['data_final']
    personalizado = True

    print(data_inicio, data_fim)

    from scripts.confere_margem_decathlon import main

    main(data_inicio, data_fim, personalizado)

    return redirect('index-relatorios')

def margem_centauro_personalizada(request):
    data_inicio = request.POST['data_inicial']
    data_fim = request.POST['data_final']
    personalizado = True

    print(data_inicio, data_fim)

    from scripts.confere_margem_centauro import main

    main(data_inicio, data_fim, personalizado)

    return redirect('index-relatorios')

def margem_mercadolivre_madz_personalizada(request):
    data_inicio = request.POST['data_inicial']
    data_fim = request.POST['data_final']
    personalizado = True

    print(data_inicio, data_fim)

    from trackeamento.scripts.margem_mercadolivre_madz import main

    main(data_inicio, data_fim, personalizado)

    return redirect('index-relatorios')

def margem_mercadolivre_redplace_personalizada(request):
    data_inicio = request.POST['data_inicial']
    data_fim = request.POST['data_final']
    personalizado = True

    print(data_inicio, data_fim, personalizado)

    from trackeamento.scripts.margem_mercadolivre_redplace import main

    main(data_inicio, data_fim, personalizado)

    return redirect('index-relatorios')

def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

def armazens_estoque_valor_custo_total(request):
    send_to_slack = False

    workbook = cria_planilha_armazem_valor_custo_total.main(send_to_slack)
    
    # Criar um objeto BytesIO para armazenar o arquivo Excel
    output = BytesIO()

    # Salvar o workbook no objeto BytesIO
    workbook.save(output)
    output.seek(0)
    
    # Criar uma resposta HTTP com o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=estoque_por_armazem.xlsx'
    response.write(output.getvalue())
    
    return response


def comparativo_estoque_magalu(request):
    connection = get_connection()
    conexao = pyodbc.connect(connection)

    empresa_personalizada = request.POST['empresa_personalizada']
    if empresa_personalizada == '23':
        armazem = '11' # full magalu madz
    elif empresa_personalizada == '24':
        armazem = '13' # full magalu redplace
    else:
        armazem = None

    file = request.FILES['file']
    df_mktp = pd.read_excel(file)
    df_mktp = df_mktp[['Código referência do produto', 'À venda', 'Programado']]
    df_mktp['ESTOQUE_MKTP'] = df_mktp['À venda'] + df_mktp['Programado']
    df_mktp = df_mktp.drop(columns=['À venda', 'Programado'])
    df_mktp.rename(columns={'Código referência do produto': 'SKU'}, inplace=True)

    # Produtos Materiais
    comando = f'''
    SELECT A.SKU,  B.COD_INTERNO, B.DESCRICAO, C.ESTOQUE AS ESTOQUE_ATON
    FROM ECOM_SKU A
    LEFT JOIN MATERIAIS B ON A.MATERIAL_ID = B.CODID
    LEFT JOIN ESTOQUE_MATERIAIS C ON A.MATERIAL_ID = C.MATERIAL_ID
    WHERE C.ARMAZEM = {armazem}
    AND A.ORIGEM_ID = {empresa_personalizada}
    '''

    # Carrega planilha
    df_ecom = pd.read_sql(comando, conexao)
    conexao.close()

    # Limpando valores com espaços vazios
    df_ecom['SKU'] = df_ecom['SKU'].str.strip()
    df_ecom['DESCRICAO'] = df_ecom['DESCRICAO'].str.strip()
    df_ecom['COD_INTERNO'] = df_ecom['COD_INTERNO'].str.strip()

    # Merge
    df = pd.merge(df_ecom, df_mktp, on='SKU', how='right')

    # Cria coluna de diferença do estoque
    df['ESTOQUE_MKTP'] = df['ESTOQUE_MKTP'].fillna(0)
    df['ESTOQUE_ATON'] = df['ESTOQUE_ATON'].fillna(0)
    df['DIFERENÇA'] = df['ESTOQUE_ATON'] - df['ESTOQUE_MKTP']

    import tempfile

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        temp_path = tmp.name
        
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='COMPARATIVO', index=False)
            worksheet = writer.sheets['COMPARATIVO']
            worksheet.auto_filter.ref = "A1:F1"
            worksheet.freeze_panes = 'A2'
            cor_laranja = PatternFill(start_color='F1C93B', end_color='F1C93B', fill_type='solid')
            celulas_laranja = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']
            for celula_referencia in celulas_laranja:
                celula = worksheet[celula_referencia]
                celula.fill = cor_laranja

    # Lê o arquivo temporário e cria a response
    with open(temp_path, 'rb') as excel_file:
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=comparativo_estoque.xlsx'

    # Remove o arquivo temporário
    os.unlink(temp_path)
    
    return response