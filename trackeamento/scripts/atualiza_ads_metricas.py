import requests
from pathlib import Path
import sys
from dotenv import load_dotenv
import os
from datetime import datetime
from datetime import datetime, timedelta
import pandas as pd
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from openpyxl.styles import Alignment, PatternFill
from time import sleep
import psutil

BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.append(str(BASE_DIR))

from mercadolivre.scripts.config import reflash
from trackeamento.models import MetricasAds

def main():
    ACCESS_TOKEN = reflash.refreshToken()
    header = {"Authorization": f"Bearer {ACCESS_TOKEN}"}
    
    # Obter a data de hoje e 7 dias atras
    hoje = datetime.now().date()
    ontem = hoje - timedelta(days=1)
    sete_dias_atras = hoje - timedelta(days=7)
    date_from = sete_dias_atras.strftime("%Y-%m-%d")
    date_to = ontem.strftime("%Y-%m-%d")

    print(date_from, date_to)

    # Consulta as campanhas ativas
    response = requests.get(f"https://api.mercadolibre.com/advertising/product_ads_2/campaigns/search?user_id=195279505&status=active&limit=50", headers=header).json()

    # Listas para armazenar dados das campanhas
    ids_list = []
    names_list = []
    last_updated_list = []
    date_created_list = []
    bidding_strategy_list = []
    status_list = []
    target_take_rate_list = []

    # Percorre cada campanha e extrai os dados
    for result in response['results']:
        ids_list.append(result['id'])
        names_list.append(result['name'])
        last_updated_list.append(result['last_updated'])
        date_created_list.append(result['date_created'])
        bidding_strategy_list.append(result['bidding_strategy'])
        status_list.append(result['status'])
        target_take_rate_list.append(result['target_take_rate'])
        

    # Percorre cada campanha
    for index, id in enumerate(ids_list):
        print(f"Campanha: {names_list[index]} - {index+1}/{len(ids_list)}")
        
        # Consulta os anúncios da campanha
        response = requests.get(f"https://api.mercadolibre.com/advertising/product_ads/ads/search?user_id=195279505&status=active&campaigns={id}", headers=header).json()

        # Percorre cada anúncio da campanha
        for result in response['results']:
            # Dados da campanha desse anúncio
            name_campanha = names_list[index]
            acos_campanha = target_take_rate_list[index]
    
            # Armazena os dados do anúncio
            mlb_anuncio = result['id']
            titulo = result['title']
            price = result['price']
            link = result['permalink']
            
            # Obtem as metricas do anúncio    
            response = requests.get(f"https://api.mercadolibre.com/advertising/product_ads_2/campaigns/{id}/ads/metrics?date_from={date_from}&date_to={date_to}&ids={mlb_anuncio}", headers=header).json()
            
            # Armazena as métricas do anúncio   
            clicks = response[0]['clicks']
            impressions = response[0]['impressions']
            cost = response[0]['cost']
            cpc = response[0]['cpc']
            ctr = response[0]['ctr']
            
            try:
                cvr = round(response[0]['sold_quantity_total'] / response[0]['clicks'] * 100, 1)
            except ZeroDivisionError:
                cvr = 0
                
            sold_quantity_total = response[0]['sold_quantity_total']
            amount_total = response[0]['amount_total']
            advertising_fee = response[0]['advertising_fee']
            organic_orders_quantity = response[0]['organic_orders_quantity']
            share = response[0]['share']
            
            # Armazena os dados no banco de dados
            novo_registro_ads = MetricasAds()
            novo_registro_ads.desde = date_from
            novo_registro_ads.ate = date_to
            novo_registro_ads.nome_campanha = name_campanha
            novo_registro_ads.acos_campanha = acos_campanha
            novo_registro_ads.mlb_anuncio = mlb_anuncio
            novo_registro_ads.titulo_anuncio = titulo
            novo_registro_ads.preco_anuncio = price
            novo_registro_ads.link_anuncio = link
            novo_registro_ads.clicks_anuncio = clicks
            novo_registro_ads.impressions_anuncio = impressions
            novo_registro_ads.cost_anuncio = cost
            novo_registro_ads.cpc_anuncio = cpc
            novo_registro_ads.ctr_anuncio = ctr
            novo_registro_ads.cvr_anuncio = cvr
            novo_registro_ads.sold_quantity_total_anuncio = sold_quantity_total
            novo_registro_ads.amount_total_anuncio = amount_total
            novo_registro_ads.advertising_fee_anuncio = advertising_fee
            novo_registro_ads.organic_orders_quantity_anuncio = organic_orders_quantity
            novo_registro_ads.share_anuncio = share
            novo_registro_ads.save()
            
def envia_slack():
    hoje = datetime.now().date()
    hoje_formatado = hoje.strftime("%d-%m-%Y")
    
    metricas = MetricasAds.objects.all()
    
    data = {
        'DESDE': [],
        'ATÉ': [],
        'CAMPANHA': [],
        'ACOS': [],
        'MLB': [],
        'TÍTULO': [],
        'PREÇO': [],
        'IMPRESSÕES': [],
        'CLIQUES': [],
        'CTR': [],
        'CPC': [],
        'TX CONV': [],
        'CUSTO': [],
        'VENDAS_TOTAIS': [],
        'REC_TOTAL': [],
        'ACOS2': [],
        'PED VENDA ORG': [],
        '% PUB/ORG': [],
        'LINK': []
    }
    
    for metrica in metricas:
        data['DESDE'].append(metrica.desde)
        data['ATÉ'].append(metrica.ate)
        data['CAMPANHA'].append(metrica.nome_campanha)
        data['ACOS'].append(metrica.acos_campanha)
        data['MLB'].append(metrica.mlb_anuncio)
        data['TÍTULO'].append(metrica.titulo_anuncio)
        data['PREÇO'].append(metrica.preco_anuncio)
        data['IMPRESSÕES'].append(metrica.impressions_anuncio)
        data['CLIQUES'].append(metrica.clicks_anuncio)
        data['CTR'].append(metrica.ctr_anuncio)
        data['CPC'].append(metrica.cpc_anuncio)
        data['TX CONV'].append(metrica.cvr_anuncio)
        data['CUSTO'].append(metrica.cost_anuncio)
        data['VENDAS_TOTAIS'].append(metrica.sold_quantity_total_anuncio)
        data['REC_TOTAL'].append(metrica.amount_total_anuncio)
        data['ACOS2'].append(metrica.advertising_fee_anuncio)
        data['PED VENDA ORG'].append(metrica.organic_orders_quantity_anuncio)
        data['% PUB/ORG'].append(metrica.share_anuncio)
        data['LINK'].append(metrica.link_anuncio)
        
    df = pd.DataFrame(data)
    
    # Converter as colunas "DESDE" e "ATÉ" para o formato desejado
    df['DESDE'] = pd.to_datetime(df['DESDE']).dt.strftime('%d-%m-%Y')
    df['ATÉ'] = pd.to_datetime(df['ATÉ']).dt.strftime('%d-%m-%Y')
    
    path_metricas = f'{hoje_formatado}-metricas_ads.xlsx'
    writer = pd.ExcelWriter(path_metricas, engine='openpyxl')
    df.to_excel(writer, index=False)
    worksheet = writer.sheets['Sheet1']

    # Adicionando filtros
    worksheet.auto_filter.ref = "A1:S1"
    
    # Alterando tamanho das colunas
    worksheet.column_dimensions['A'].width = '11.16'
    worksheet.column_dimensions['B'].width = '10.44'
    worksheet.column_dimensions['C'].width = '17.59'
    worksheet.column_dimensions['D'].width = '10.44'
    worksheet.column_dimensions['E'].width = '14.87'
    worksheet.column_dimensions['F'].width = '59.44'
    worksheet.column_dimensions['G'].width = '11.44'
    worksheet.column_dimensions['H'].width = '16.73'
    worksheet.column_dimensions['I'].width = '13.02'
    worksheet.column_dimensions['J'].width = '8.87'
    worksheet.column_dimensions['K'].width = '9.02'
    worksheet.column_dimensions['L'].width = '13.44'
    worksheet.column_dimensions['M'].width = '10.86'
    worksheet.column_dimensions['N'].width = '20.3'
    worksheet.column_dimensions['O'].width = '15.44'
    worksheet.column_dimensions['P'].width = '11.44'
    worksheet.column_dimensions['Q'].width = '20.3'
    worksheet.column_dimensions['R'].width = '16.26'
    worksheet.column_dimensions['S'].width = '10.02'
    
    # Congelando painel
    worksheet.freeze_panes = 'A2'
    
    # Define a formatação de centralização
    alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
            
    # Atualiza cores do cabeçalho
    cor_header = 'F79646'
    for i, row in enumerate(worksheet.iter_rows(min_row=1, max_row=1)):
        if i == 0:
            fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type='solid')
        for cell in row:
            cell.fill = fill
            
    # Adicionar o cifrão brasileiro às células das colunas
    columns = ['G', 'K', 'M', 'O', 'P']
    for col_letter in columns:
        for cell in worksheet[col_letter][1:]:
            cell.number_format = 'R$ #,##0.00'
            
    # Definir o formato de porcentagem para as colunas
    columns = ['J', 'L', 'R']
    for col_letter in columns:
        for cell in worksheet[col_letter][1:]:
            cell.number_format = '0.0%'
            cell.value /= 100
            
    writer._save()
        
    # Envia slack
    load_dotenv()

    client = WebClient(token=os.environ['SLACK_BOT_TOKEN'])
    SLACK_CHANNEL_ID='C030X3UMR3M'
    #SLACK_CHANNEL_ID='C045HEE4G7L'
    
    message = f'MERCADO ADS! :money_mouth_face:'
    
    # Send message
    try:
        client.chat_postMessage(channel=SLACK_CHANNEL_ID, text=message)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    # Send file
    try:
        client.files_upload_v2(channel=SLACK_CHANNEL_ID, file=path_metricas, filename=path_metricas)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    # Remove arquivo
    try:
        os.remove(path_metricas)
    except Exception as e:
        print(e)