import requests
from pathlib import Path
import sys
from dotenv import load_dotenv
import pyodbc
import os
from datetime import datetime, timedelta, timezone
import pandas as pd
from slack_sdk import WebClient
from openpyxl.styles import NamedStyle
from slack_sdk.errors import SlackApiError
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import json
from time import sleep


BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.append(str(BASE_DIR))

from mercadolivre.scripts.configFitgear import reflash
from scripts.connect_to_database import get_connection

def main(inicio_data_personalizada, fim_data_personalizada, personalizado):
    ACCESS_TOKEN = reflash.refreshToken()
    header = {"Authorization": f"Bearer {ACCESS_TOKEN}"}

    list_id = []
    list_sale_fee = []
    list_pack_id = []
    list_unit_price = []
    list_paid_amount = []
    list_full_unit_price = []
    list_date_created = []
    list_mlb = []
    list_coupon_amount_divided = []
    list_shipping_cost_buyer = []
    list_imposto_dez_porcento_shipping_cost_buyer = []
    list_title = []
    list_category = []
    list_quantity_unit = []
    list_listing_type_id = []
    list_hiperlink = []
    list_campanha_bool = []
    list_imposto = []
    list_operacao = []

    # Cria a data
    # Define o timezone como UTC
    if personalizado == False:
        print('falso')
        hoje = datetime.now(timezone.utc).date()
        debito_dias = hoje - timedelta(days=1)
        date_from_strf = debito_dias.strftime("%Y-%m-%dT00:00:00.000Z")
        link_get_total_orders = f"https://api.mercadolibre.com/orders/search?seller=429996719&order.status=paid&order.date_created.from={date_from_strf}&sort=date_desc"
    else:
        print('true')
        # Converte strings para objetos de data, se necessário
        inicio_data = datetime.strptime(inicio_data_personalizada, "%Y-%m-%d")
        date_from_strf = inicio_data.strftime("%Y-%m-%dT00:00:00.000Z")
        
        fim_data = datetime.strptime(fim_data_personalizada, "%Y-%m-%d")
        fim_data = fim_data + timedelta(days=1)
        date_to_strf = fim_data.strftime("%Y-%m-%dT00:00:00.000Z")
        link_get_total_orders = f"https://api.mercadolibre.com/orders/search?seller=429996719&order.status=paid&order.date_created.from={date_from_strf}&order.date_created.to={date_to_strf}&sort=date_desc"


    response = requests.get(link_get_total_orders, headers=header)
    error_exit = 0
    while response.status_code == 500 or response.status_code == 403:
        error_exit += 1
        if error_exit > 6:
            print('ERROR 500 or 403')
            break
        sleep(10)
        print('Retrying 1...')
        response = requests.get(link_get_total_orders, headers=header)
    response = response.json()

    total_orders = response['paging']['total']
    print(f'Total de pedidos: {total_orders}')

    offset = 0
    # Obtem todos os pedidos
    while offset < total_orders:
        
        # Consulta pedido
        for result in response['results']:
            
            # Settings variables
            quantity_itens = len(result['order_items'])
    
            # shipping_cost cliente
            shipping_id = result['shipping']['id']
            response_shipping = requests.get(f"https://api.mercadolibre.com/shipments/{shipping_id}", headers=header).json()
            shipping_cost_buyer = response_shipping['shipping_option']['cost']

            # Consulta itens
            for item in result['order_items']:
                
                # Geral
                id_order = result['id']
                pack_id = result['pack_id']
                date_created = result['date_created']
                paid_amount = result['paid_amount']
                coupon_amount = result['coupon']['amount']
                coupon_amount_divided = coupon_amount / quantity_itens
                

                # Old
                # shipping_cost_buyer = 0
                # for payments in result['payments']:
                #     shipping_cost_buyer += payments['shipping_cost']

                # calcula 10 porcento de shipping cost
                imposto_dez_porcento_shipping_cost_buyer = (shipping_cost_buyer * 0.1)
                list_imposto_dez_porcento_shipping_cost_buyer.append(imposto_dez_porcento_shipping_cost_buyer)

                # Cria hiperlink
                hiperlink = f'https://www.mercadolivre.com.br/vendas/{id_order}/detalhe'
                list_hiperlink.append(hiperlink)
                
                # Item especifico
                sale_fee = item['sale_fee']
                if sale_fee == None:
                    sale_fee = 0
                else:
                    sale_fee = sale_fee * quantity_itens

                mlb = item['item']['id']
                title = item['item']['title']
                category = item['item']['category_id']
                quantity = item['quantity']
                unit_price = item['unit_price']
                full_unit_price = item['full_unit_price']
                listing_type_id = item['listing_type_id']

                # Verifica se está em campanha
                if unit_price == full_unit_price:
                    list_campanha_bool.append('NÃO')
                else:
                    list_campanha_bool.append('SIM')

                # Imposto e operação - 10% de unit_price
                imposto = (unit_price * 0.1)
                operacao = (unit_price * 0.1)
                list_imposto.append(imposto)
                list_operacao.append(operacao)

                list_id.append(id_order)
                list_sale_fee.append(sale_fee)
                list_pack_id.append(pack_id)
                list_unit_price.append(unit_price)
                list_full_unit_price.append(full_unit_price)
                list_date_created.append(date_created)
                list_mlb.append(mlb)
                list_paid_amount.append(paid_amount)
                list_coupon_amount_divided.append(coupon_amount_divided)
                list_shipping_cost_buyer.append(shipping_cost_buyer)
                list_title.append(title)
                list_category.append(category)
                list_quantity_unit.append(quantity)
                list_listing_type_id.append(listing_type_id)


        # Incrementa o offset
        offset += 51

        if personalizado == False:
            link_get_total_orders_offset = f"https://api.mercadolibre.com/orders/search?seller=429996719&order.status=paid&order.date_created.from={date_from_strf}&offset={offset}&sort=date_desc"
        else:
            # Converte strings para objetos de data
            inicio_data = datetime.strptime(inicio_data_personalizada, "%Y-%m-%d")
            fim_data = datetime.strptime(fim_data_personalizada, "%Y-%m-%d")
            date_from_strf = inicio_data.strftime("%Y-%m-%dT00:00:00.000Z")
            date_to_strf = fim_data.strftime("%Y-%m-%dT00:00:00.000Z")
            
            link_get_total_orders_offset = f"https://api.mercadolibre.com/orders/search?seller=429996719&order.status=paid&order.date_created.from={date_from_strf}&order.date_created.to={date_to_strf}&offset={offset}&sort=date_desc"

        response = requests.get(link_get_total_orders_offset, headers=header)
        while response.status_code == 500 or response.status_code == 403:
            sleep(10)
            print('Retrying 2...')
            print(response.text)
            response = requests.get(link_get_total_orders_offset, headers=header)
        response = response.json()

    
    # Transforma em lista de strings
    list_id = [str(i) for i in list_id]
    list_pack_id = [str(i) for i in list_pack_id]

    # Replace palavra None por vazio em list_pack_id
    list_pack_id = [i.replace('None', '') for i in list_pack_id]

    # Adiciona as listas em DF
    df = pd.DataFrame({
                       'DATA': list_date_created,
                       'ID': list_id, 
                       'PACK_ID': list_pack_id, 
                       'MLB': list_mlb,
                       'TITULO': list_title,
                       'QUANTIDADE': list_quantity_unit,
                       'CATEGORIA': list_category,
                       'TIPO ANUNCIO': list_listing_type_id,
                       'PREÇO': list_full_unit_price,
                       'PREÇO CAMPANHA': list_unit_price,
                       'IMPOSTO': list_imposto,
                       'OPERAÇÃO': list_operacao,
                       'CAMPANHA': list_campanha_bool,
                       'VLR TOTAL NOTA': list_paid_amount,
                       'TAXA': list_sale_fee, 
                       'CUPOM': list_coupon_amount_divided,
                       'FRETE': list_shipping_cost_buyer,
                       'FRETE 10%': list_imposto_dez_porcento_shipping_cost_buyer,
                       'LINK': list_hiperlink
                       })
    

    # Format the DATA column to the desired format
    df['DATA'] = pd.to_datetime(df['DATA'])
    df['DATA'] = df['DATA'].dt.strftime('%d-%m-%Y %H:%M:%S')

    # Renomeia valores TIPO ANUNCIO
    df.loc[df['TIPO ANUNCIO'] == 'gold_special', 'TIPO ANUNCIO'] = 'CLÁSSICO'
    df.loc[df['TIPO ANUNCIO'] == 'gold_pro', 'TIPO ANUNCIO'] = 'PREMIUM'

    # Planilha aton
    connection = get_connection()
    conexao = pyodbc.connect(connection)
    cursor = conexao.cursor()

    comando = f'''
    SELECT B.CODID, B.COD_INTERNO, B.DESCRICAO, B.VLR_CUSTO, A.SKU AS MLB
    FROM ECOM_SKU A
    LEFT JOIN MATERIAIS B ON A.MATERIAL_ID = B.CODID
    LEFT JOIN ECOM_ORIGEM C ON A.ORIGEM_ID = C.ORIGEM_ID
    WHERE C.API = 'Mercado Livre'
    ORDER BY CODID
    '''

    df_aton = pd.read_sql(comando, conexao)

    # Limpa valores com espaços vazios
    df_aton['MLB'] = df_aton['MLB'].str.strip()
    df_aton['COD_INTERNO'] = df_aton['COD_INTERNO'].str.strip()
    df_aton['DESCRICAO'] = df_aton['DESCRICAO'].str.strip()

    # fecha conexão
    cursor.close()
    conexao.close()

    df_aton_unique = df_aton.drop_duplicates(subset=['MLB'])
    df_completo = pd.merge(df, df_aton_unique, on=['MLB'], how='left')

    # lucro liq % é preço - imposto - opepraçã
    df_completo['LUCRO LIQ $'] = df_completo['PREÇO CAMPANHA'] - df_completo['IMPOSTO'] - df_completo['OPERAÇÃO'] - df_completo['TAXA'] - df_completo['VLR_CUSTO'] - df_completo['FRETE 10%']
    df_completo['LUCRO LIQ %'] = round((df_completo['LUCRO LIQ $'] / df_completo['PREÇO CAMPANHA']) * 100, 2)

    # altera ordem das colunas
    df_completo = df_completo[['CODID', 'COD_INTERNO', 'ID', 'PACK_ID', 'MLB', 'DESCRICAO', 'TITULO', 'VLR_CUSTO', 'PREÇO', 'PREÇO CAMPANHA', 'CAMPANHA', 'VLR TOTAL NOTA',
                                'FRETE', 'FRETE 10%','DATA', 'QUANTIDADE', 'CATEGORIA', 'TIPO ANUNCIO', 'IMPOSTO', 'OPERAÇÃO', 'TAXA', 'CUPOM', 'LUCRO LIQ $', 'LUCRO LIQ %', 'LINK']]

    # Exporta em excel
    if personalizado == False:
        name_file_excel = 'margem_mercadolivre_' + str(hoje) + '.xlsx'
    else:
        name_file_excel = 'margem_mercadolivrepersonalizado.xlsx'
    writer = pd.ExcelWriter(name_file_excel, engine='openpyxl')
    df_completo.to_excel(writer, sheet_name='MERCADOLIVRE', index=False)
    worksheet = writer.sheets['MERCADOLIVRE']

    # Adicionando filtros
    worksheet.auto_filter.ref = "A1:Y1"

    # Congelando painel
    worksheet.freeze_panes = 'A2'

    # Definir a cor
    cor_laranja = PatternFill(start_color='F1C93B', end_color='F1C93B', fill_type='solid')
    cor_verde = PatternFill(start_color='96C291', end_color='96C291', fill_type='solid')
    cor_branco = PatternFill(start_color='F4EEEE', end_color='F4EEEE', fill_type='solid')

    # Lista Porcentagem laranja
    celulas_laranja = ['X1']
    for celula_referencia in celulas_laranja:
        celula = worksheet[celula_referencia]
        celula.fill = cor_laranja
        
    # Lista Porcentagem Verde
    celulas_verde = ['H1', 'I1', 'J1', 'L1', 'M1', 'N1', 'S1', 'T1', 'U1', 'V1', 'W1']
    for celula_referencia in celulas_verde:
        celula = worksheet[celula_referencia]
        celula.fill = cor_verde

    # Lista Porcentagem Branco
    celulas_branco = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'K1', 'O1', 'P1', 'Q1', 'R1', 'Y1']
    for celula_referencia in celulas_branco:
        celula = worksheet[celula_referencia]
        celula.fill = cor_branco

    # Transforma coluna Y em clicavel no excel
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=25)  # Coluna Y é a 25ª coluna (1-indexed)
        link = cell.value
        if link:
            cell.value = f'=HYPERLINK("{link}", "{link}")'

    # Ajustar o tamanho das colunas automaticamente
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width


    # Aplicar a formatação condicional na coluna X (coluna 24, 1-indexed)
    cor_vermelha = PatternFill(start_color='FA8072', end_color='FA8072', fill_type='solid')
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=24)  # Coluna X é a 24ª coluna (1-indexed)
        if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 10:
            cell.fill = cor_vermelha


    writer._save()
            
    # Envia slack
    load_dotenv()

    client = WebClient(token=os.environ['SLACK_BOT_TOKEN'])
    SLACK_CHANNEL_ID='C045HEE4G7L'

    message = f'MERCADO LIVRE MARGEM! :heavy_division_sign:'

    # Send message
    try:
        client.chat_postMessage(channel=SLACK_CHANNEL_ID, text=message)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    # Send file
    try:
        client.files_upload_v2(channel=SLACK_CHANNEL_ID, file=name_file_excel, filename=name_file_excel)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    writer.close()
        
    # Remove arquivo
    try:
        os.remove(name_file_excel)
    except Exception as e:
        print(e)