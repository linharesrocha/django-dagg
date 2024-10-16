from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import PosicaoNetshoes, MetricasMercadoLivre, MatchNetshoes, NetshoesTarifas, DecathlonTarifas, CentauroTarifas
from django.db.models import OuterRef, Subquery, Max
import pandas as pd
from datetime import datetime
from io import BytesIO
from trackeamento.scripts import atualiza_netshoes, atualiza_mercadolivre, verifica_match_netshoes
from openpyxl.styles import Alignment, PatternFill
from django.contrib.messages import constants
from django.contrib import messages
import unidecode
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import seaborn as sns
from io import BytesIO
import base64

def index(request):
    return render(request, 'trackeamento/index.html')

def match_netshoes(request):
    return render(request, 'trackeamento/match_netshoes/match-netshoes.html')

def ajuste_tarifas(request):
    return render(request, 'trackeamento/ajuste_tarifas/ajuste_tarifas.html')

def painel_match_netshoes(request):
    ultimos_valores = MatchNetshoes.objects.filter(
        id__in=Subquery(
            MatchNetshoes.objects.filter(
                sku_match=OuterRef('sku_match')
            ).values('sku_match').annotate(
                ultima_id=Max('id')
            ).values('ultima_id')
        )
    )

    for valor in ultimos_valores:
        valor.ultima_atualizacao = valor.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
    
    match_netshoes = {
        'match_netshoes': ultimos_valores
    }

    return render(request, 'trackeamento/match_netshoes/painel-match-netshoes.html', match_netshoes)

def posicao_netshoes(request):
    return render(request, 'trackeamento/netshoes/posicao-netshoes.html')


def painel_posicao_netshoes(request):

    ultimos_valores = PosicaoNetshoes.objects.filter(
        id__in=Subquery(
            PosicaoNetshoes.objects.filter(
                sku_netshoes=OuterRef('sku_netshoes')
            ).values('sku_netshoes').annotate(
                ultima_id=Max('id')
            ).values('ultima_id')
        )
    )
    
    for valor in ultimos_valores:
        valor.ultima_atualizacao = valor.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')
    
    posicoes_netshoes = {
        'posicoes_netshoes': ultimos_valores
    }

    return render(request, 'trackeamento/netshoes/painel-posicao-netshoes.html', posicoes_netshoes)


def cadastrar_posicao_netshoes(request):
    termo = request.POST['termo-cadastro']
    termo = unidecode.unidecode(termo)
    sku_netshoes = request.POST['sku-netshoes-cadastro']
    nome = request.POST['nome-cadastro']
    tipo_anuncio = request.POST['tipo_anuncio']
    canal = request.POST['canal']
    
    # Verifica campo vazio
    if termo == '' or sku_netshoes == '' or nome == '' or canal == '':
        messages.add_message(request, constants.ERROR, 'Preencha todos os campos!')
        return redirect('posicao-netshoes')
    
    # Transforma em lower
    termo = termo.lower().strip()
    sku_netshoes = sku_netshoes.upper().strip()
    nome = nome.strip()
    canal = canal.upper()
    
    # Verifica se existe no banco
    my_obj = PosicaoNetshoes.objects.filter(sku_netshoes=sku_netshoes).first()
    if my_obj is not None:
        messages.add_message(request, constants.ERROR, 'SKU já existe no banco de dados!')
        return redirect('posicao-netshoes')
    
    # Cadastra
    nova_posicao = PosicaoNetshoes(
        nome=nome,
        termo_busca=termo,
        sku_netshoes=sku_netshoes,
        anuncio_concorrente=tipo_anuncio == 'concorrente',
        canal=canal
    )
    nova_posicao.save()
    
    messages.add_message(request, constants.SUCCESS, 'Cadastrado com sucesso!')
    return redirect('posicao-netshoes')

def cadastrar_match_netshoes(request):
    titulo = request.POST['titulo-cadastro']
    titulo = unidecode.unidecode(titulo)
    sku_match = request.POST['sku-match-netshoes-cadastro']
    nome_loja = request.POST['nome-loja']
    
    # Verifica campo vazio
    if titulo == '' or sku_match == '' or nome_loja == '':
        messages.add_message(request, constants.ERROR, 'Preencha os campos vazios!')
        return redirect('match-netshoes')
    
    # Transforma em lower
    sku_match = sku_match.upper()
    sku_match = sku_match.strip()
    
    # Verifica se existe no banco
    my_obj = MatchNetshoes.objects.filter(sku_match=sku_match).first()
    if my_obj is not None:
        messages.add_message(request, constants.ERROR, 'SKU já existe no banco de dados!')
        return redirect('match-netshoes')
    
    # Cadastra
    nova_posicao = MatchNetshoes()
    nova_posicao.titulo_produto = titulo
    nova_posicao.sku_match = sku_match
    nova_posicao.nome_loja = nome_loja
    nova_posicao.status = None   
    nova_posicao.save()
    
    messages.add_message(request, constants.SUCCESS, 'Cadastrado com sucesso!')
    return redirect('match-netshoes')


def remover_posicao_netshoes(request):
    sku_netshoes = request.POST['sku-netshoes-delete']
    
    # Verifica campo vazio
    if sku_netshoes == '':
        messages.add_message(request, constants.ERROR, 'Preencha o campo vazio!')
        return redirect('posicao-netshoes')
    
    # Verifica se existe no banco
    my_obj = PosicaoNetshoes.objects.filter(sku_netshoes=sku_netshoes).first()
    if my_obj is None:
        messages.add_message(request, constants.ERROR, 'SKU não existe no banco de dados!')
        return redirect('posicao-netshoes')
    
    # Deleta
    PosicaoNetshoes.objects.filter(sku_netshoes=sku_netshoes).delete()

    messages.add_message(request, constants.SUCCESS, 'Removido com sucesso!')
    return redirect('posicao-netshoes')

def remover_match_netshoes(request):
    sku_match = request.POST['sku-netshoes-delete']
    
    # Verifica campo vazio
    if sku_match == '':
        messages.add_message(request, constants.ERROR, 'Preencha o campo vazio!')
        return redirect('match-netshoes')
    
    # Verifica se existe no banco
    my_obj = MatchNetshoes.objects.filter(sku_match=sku_match).first()
    if my_obj is None:
        messages.add_message(request, constants.ERROR, 'SKU não existe no banco de dados!')
        return redirect('match-netshoes')
    
    # Deleta
    MatchNetshoes.objects.filter(sku_match=sku_match).delete()

    messages.add_message(request, constants.SUCCESS, 'Removido com sucesso!')
    return redirect('match-netshoes')


def baixar_historico(request):
    posicoes_netshoes = PosicaoNetshoes.objects.all()
    
    if len(posicoes_netshoes) == 0:
        messages.add_message(request, constants.ERROR, 'Não há dados para baixar!')
        return redirect('painel-posicao-netshoes')
     
    dados = {
         'id': [posicao_netshoes.id for posicao_netshoes in posicoes_netshoes],
         'pagina': [posicao_netshoes.pagina for posicao_netshoes in posicoes_netshoes],
         'posicao': [posicao_netshoes.posicao for posicao_netshoes in posicoes_netshoes],
         'nome': [posicao_netshoes.nome for posicao_netshoes in posicoes_netshoes],
         'sku_netshoes': [posicao_netshoes.sku_netshoes for posicao_netshoes in posicoes_netshoes],
         'termo_busca': [posicao_netshoes.termo_busca for posicao_netshoes in posicoes_netshoes],
         'variacao': [posicao_netshoes.variacao for posicao_netshoes in posicoes_netshoes],
         'anuncio_concorrente': [posicao_netshoes.anuncio_concorrente for posicao_netshoes in posicoes_netshoes],
         'ultima_atualizacao': [posicao_netshoes.ultima_atualizacao for posicao_netshoes in posicoes_netshoes]
     }
     
    df = pd.DataFrame(dados)
    
    df['ultima_atualizacao'] = df['ultima_atualizacao'].dt.tz_localize(None)
    df['pagina'].fillna('None', inplace=True)
    df['variacao'].fillna('None', inplace=True)
    df['posicao'].fillna('None', inplace=True)
    df['ultima_atualizacao'] = pd.to_datetime(df['ultima_atualizacao'])
    df['ultima_atualizacao'] = df['ultima_atualizacao'].dt.strftime('%d-%m-%Y %H:%M:%S')
    
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name='HISTORICO',index=False)
    worksheet = writer.sheets['HISTORICO']
    
    # Adicionando filtros
    worksheet.auto_filter.ref = "A1:I1"
    
    # Alterando tamanho das colunas
    worksheet.column_dimensions['A'].width = '3.30'
    worksheet.column_dimensions['B'].width = '10.71'
    worksheet.column_dimensions['C'].width = '11.57'
    worksheet.column_dimensions['D'].width = '51'
    worksheet.column_dimensions['E'].width = '15.57'
    worksheet.column_dimensions['F'].width = '16.29'
    worksheet.column_dimensions['G'].width = '11.14'
    worksheet.column_dimensions['H'].width = '19.29'
    worksheet.column_dimensions['I'].width = '21.71'
        
    # Escondendo coluna ID
    worksheet.column_dimensions['A'].hidden = True
    
    # Congelando painel
    worksheet.freeze_panes = 'A2'
    
    # Define a formatação de centralização
    alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Atualiza cores do cabeçalho
    cor_header = 'F79646'
    for i, row in enumerate(worksheet.iter_rows(min_row=1, max_row=1)):
        if i == 0:
            fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type='solid')
        for cell in row:
            cell.fill = fill
    
    writer._save()
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=historico_netshoes_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
    return response


def atualizar_historico(request):
    slack = False
    atualiza_netshoes.main(slack)
    
    messages.add_message(request, constants.SUCCESS, 'Histórico Atualizado!')
    return redirect('painel-posicao-netshoes')


def gerar_grafico(sku_netshoes):
    itens = PosicaoNetshoes.objects.filter(sku_netshoes=sku_netshoes).all()

    # Extrai as colunas de DATA e PAGINA dos dados 
    datas = [i.ultima_atualizacao.strftime('%d-%m') for i in itens]
    paginas = [i.pagina for i in itens]
    paginas = [0 if pagina is None else pagina for pagina in paginas]

    plt.clf()
    sns.set(style="whitegrid")
    sns.lineplot(x=datas, y=paginas)
    plt.yticks(range(int(min(paginas)), int(max(paginas))+1))
    sns.set(rc={'figure.figsize':(50, 50)})
    plt.xlabel('Data')
    plt.ylabel('Página')
    plt.title('Histórico de Páginas')
    plt.xticks(rotation=50)
    
    # Cria um buffer de memória para salvar o gráfico
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    imagem_grafico = base64.b64encode(buffer.getvalue()).decode('utf-8')
    buffer.close()

    return imagem_grafico

def item_posicao_netshoes(request, sku_netshoes):
    item = PosicaoNetshoes.objects.filter(sku_netshoes=sku_netshoes).first()
    
    imagem_grafico = gerar_grafico(sku_netshoes)
    
    return render(request, 'trackeamento/netshoes/item-posicao-netshoes.html', {'item': item, 'imagem_grafico': imagem_grafico})


def item_alterar_nome(request):
    novo_nome = request.POST['novo-nome']
    sku_netshoes = request.POST.get('sku_netshoes')
    PosicaoNetshoes.objects.filter(sku_netshoes=sku_netshoes).update(nome=novo_nome)

    return redirect('painel-posicao-netshoes')


def metricas_mercadolivre(request):
    # Caso o usuário entre
    if request.method == 'GET':
        return render(request, 'trackeamento/mercadolivre/metricas-mercadolivre.html')
    
    # Caso o usário faça alguma ação
    elif request.method == 'POST':
        action = request.POST.get('action')
       
       # Caso a ação seja cadastrar um MLB
        if action == 'cadastrar':
            
            # Tratando os dados
            termo = request.POST['termo-cadastro'].lower().strip()
            termo = unidecode.unidecode(termo)
            mlb = request.POST['mlb-mercadolivre-cadastro'].upper().replace('-','').strip()
            
            if not mlb.startswith('MLB'):
                messages.add_message(request, constants.ERROR, 'MLB não está começando com MLB!')
                return redirect('metricas-mercadolivre')
            
            # Verifica se existe no banco
            objects = MetricasMercadoLivre.objects.filter(mlb_anuncio=mlb)
            if objects is not None:
                # Percorre cada objecto e verifica se o termo é igual
                for obj in objects:
                    if termo == obj.termo_busca:
                        messages.add_message(request, constants.ERROR, 'MLB já está cadastrado com o mesmo termo!')
                        return redirect('metricas-mercadolivre')
            
            # Cadastra
            MetricasMercadoLivre(termo_busca=termo, mlb_anuncio=mlb).save()
            
            messages.add_message(request, constants.SUCCESS, 'Cadastrado com sucesso!')
            return redirect('metricas-mercadolivre')
        
        # Caso a ação seja remover um MLB
        elif action == 'remover':
            mlb_anuncio = request.POST['mlb-mercadolivre-delete']
            termo = request.POST['termo-mercadolivre-delete']
            
            # Verifica se existe no banco de dados
            my_obj = MetricasMercadoLivre.objects.filter(mlb_anuncio=mlb_anuncio).filter(termo_busca=termo).first()
            if my_obj is None:
                messages.add_message(request, constants.ERROR, 'MLB ou Termo não existe no banco de dados!')
                return redirect('metricas-mercadolivre')
            # Caso exista, deleta
            else:
                MetricasMercadoLivre.objects.filter(mlb_anuncio=mlb_anuncio).filter(termo_busca=termo).delete()
                messages.add_message(request, constants.SUCCESS, 'Removido com sucesso!')
                return redirect('metricas-mercadolivre')
        
        return render(request, 'trackeamento/mercadolivre/metricas-mercadolivre.html')
    

def atualizar_metricas_mercadolivre(request):
    slack = False
    atualiza_mercadolivre.main(slack)
    
    messages.add_message(request, constants.SUCCESS, 'Métricas Atualizadas!')
    return redirect('metricas-mercadolivre')


def baixar_historico_mercadolivre(request):
    lista_mercadolivre = MetricasMercadoLivre.objects.all()
    
    if len(lista_mercadolivre) == 0:
        messages.add_message(request, constants.ERROR, 'Não há dados para baixar!')
        return redirect('metricas-mercadolivre')
     
    dados = {
        'id': [anuncio.id for anuncio in lista_mercadolivre],
        'nome': [anuncio.nome for anuncio in lista_mercadolivre],
        'termo_busca': [anuncio.termo_busca for anuncio in lista_mercadolivre],
        'mlb_anuncio': [anuncio.mlb_anuncio for anuncio in lista_mercadolivre],
        'posicao': [anuncio.posicao for anuncio in lista_mercadolivre],
        'pagina': [anuncio.pagina for anuncio in lista_mercadolivre],
        'posicao_full': [anuncio.posicao_full for anuncio in lista_mercadolivre],
        'pagina_full': [anuncio.pagina_full for anuncio in lista_mercadolivre],
        'visita_diaria': [anuncio.visita_diaria for anuncio in lista_mercadolivre],
        'visita_total': [anuncio.visita_total for anuncio in lista_mercadolivre],
        'vendas_diaria': [anuncio.vendas_diaria for anuncio in lista_mercadolivre],
        'vendas_total': [anuncio.vendas_total for anuncio in lista_mercadolivre],
        'vende_a_cada_visita': [anuncio.vende_a_cada_visita for anuncio in lista_mercadolivre],
        'taxa_conversao_diaria': [anuncio.taxa_conversao_diaria for anuncio in lista_mercadolivre],
        'taxa_conversao_total': [anuncio.taxa_conversao_total for anuncio in lista_mercadolivre],
        'pontuacao_anuncio': [anuncio.pontuacao_anuncio for anuncio in lista_mercadolivre],
        'criacao_anuncio': [anuncio.criacao_anuncio for anuncio in lista_mercadolivre],
        'ultima_atualizacao': [anuncio.ultima_atualizacao for anuncio in lista_mercadolivre]
    }
     
    df = pd.DataFrame(dados)
    
    df['ultima_atualizacao'] = df['ultima_atualizacao'].dt.tz_localize(None)
    df['nome'].fillna('None', inplace=True)
    df['posicao'].fillna('None', inplace=True)
    df['pagina'].fillna('None', inplace=True)
    df['posicao_full'].fillna('None', inplace=True)
    df['pagina_full'].fillna('None', inplace=True)
    df['visita_diaria'].fillna('None', inplace=True)
    df['visita_total'].fillna('None', inplace=True)
    df['vendas_diaria'].fillna('None', inplace=True)
    df['vendas_total'].fillna('None', inplace=True)
    df['vende_a_cada_visita'].fillna('None', inplace=True)
    df['taxa_conversao_diaria'].fillna('None', inplace=True)
    df['taxa_conversao_total'].fillna('None', inplace=True)
    df['pontuacao_anuncio'].fillna('None', inplace=True)
    df['criacao_anuncio'].fillna('None', inplace=True)
    df['ultima_atualizacao'] = pd.to_datetime(df['ultima_atualizacao'])
    df['ultima_atualizacao'] = df['ultima_atualizacao'].dt.strftime('%d-%m-%Y %H:%M:%S')
    
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name='HISTORICO',index=False)
    worksheet = writer.sheets['HISTORICO']
    
    # Adicionando filtros
    primeira_coluna = worksheet.min_column
    ultima_coluna = worksheet.max_column
    worksheet.auto_filter.ref = f'{get_column_letter(primeira_coluna)}1:{get_column_letter(ultima_coluna)}1'
    
    # Alterando tamanho das colunas
    worksheet.column_dimensions['A'].width = '6.57'
    worksheet.column_dimensions['B'].width = '59'
    worksheet.column_dimensions['C'].width = '17.86'
    worksheet.column_dimensions['D'].width = '15.57'
    worksheet.column_dimensions['E'].width = '10.86'
    worksheet.column_dimensions['F'].width = '10.71'
    worksheet.column_dimensions['G'].width = '15.57'
    worksheet.column_dimensions['H'].width = '14.71'
    worksheet.column_dimensions['I'].width = '15.57'
    worksheet.column_dimensions['J'].width = '14.71'
    worksheet.column_dimensions['K'].width = '17.29'
    worksheet.column_dimensions['L'].width = '16.29'
    worksheet.column_dimensions['M'].width = '23.43'
    worksheet.column_dimensions['N'].width = '24.86'
    worksheet.column_dimensions['O'].width = '24'
    worksheet.column_dimensions['P'].width = '22.43'
    worksheet.column_dimensions['Q'].width = '19.14'
    worksheet.column_dimensions['R'].width = '21.71'

    # Escondendo coluna ID
    worksheet.column_dimensions['A'].hidden = True
    
    # Congelando painel
    worksheet.freeze_panes = 'A2'
    
    # Define a formatação de centralização
    alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
            
    # Atualiza cores do cabeçalho
    cor_header = 'F79646'
    for i, row in enumerate(worksheet.iter_rows(min_row=1, max_row=1)):
        if i == 0:
            fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type='solid')
        for cell in row:
            cell.fill = fill
    
    writer._save()
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=historico_mercadolivre_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
    return response


def ajuste_tarifas_visualizar(request):
    netshoes_tarifas = NetshoesTarifas.objects.filter(id=1).first()
    centauro_tarifas = CentauroTarifas.objects.filter(id=1).first()
    decathlon_tarifas = DecathlonTarifas.objects.filter(id=1).first()

    return render(request, 'trackeamento/ajuste_tarifas/ajuste_tarifas.html', {'netshoes_tarifas': netshoes_tarifas, 'centauro_tarifas': centauro_tarifas, 'decathlon_tarifas': decathlon_tarifas})

def atualizar_tarifa_netshoes(request):
    netshoes_tarifas = NetshoesTarifas.objects.filter(id=1).first()

    if request.POST['dagg-comissao'].isnumeric():
        netshoes_tarifas.dagg_comissao_padrao = request.POST['dagg-comissao']

    if request.POST['red-comissao'].isnumeric():
        netshoes_tarifas.red_comissao_padrao = request.POST['red-comissao']

    if request.POST['pisste-comissao'].isnumeric():
        netshoes_tarifas.pisste_comissao_padrao = request.POST['pisste-comissao']

    if request.POST['tarifa-fixa'].isnumeric():
        netshoes_tarifas.tarifa_fixa = request.POST['tarifa-fixa']

    if request.POST['operacao'].isnumeric():
        netshoes_tarifas.operacao = request.POST['operacao']

    if request.POST['imposto'].isnumeric():
        netshoes_tarifas.imposto = request.POST['imposto']

    netshoes_tarifas.save()

    messages.add_message(request, constants.SUCCESS, 'Tarifas Atualizadas!')
    return redirect('/trackeamento/ajuste-tarifas')



def atualizar_tarifa_centauro(request):
    centauro_tarifas = CentauroTarifas.objects.filter(id=1).first()

    if request.POST['dagg-comissao'].isnumeric():
        centauro_tarifas.dagg_comissao_padrao = request.POST['dagg-comissao']

    if request.POST['operacao'].isnumeric():
        centauro_tarifas.operacao = request.POST['operacao']
    
    if request.POST['imposto'].isnumeric():
        centauro_tarifas.imposto = request.POST['imposto']
    
    centauro_tarifas.save()

    messages.add_message(request, constants.SUCCESS, 'Tarifas Atualizadas!')
    return redirect('/trackeamento/ajuste-tarifas')


def atualizar_tarifa_decathlon(request):
    decathlon_tarifas = DecathlonTarifas.objects.filter(id=1).first()

    if request.POST['dagg-comissao'].isnumeric():
        decathlon_tarifas.dagg_comissao_padrao = request.POST['dagg-comissao']

    if request.POST['operacao'].isnumeric():
        decathlon_tarifas.operacao = request.POST['operacao']
    
    if request.POST['imposto'].isnumeric():
        decathlon_tarifas.imposto = request.POST['imposto']
    
    decathlon_tarifas.save()

    messages.add_message(request, constants.SUCCESS, 'Tarifas Atualizadas!')
    return redirect('/trackeamento/ajuste-tarifas')