import pyodbc
import pandas as pd
import warnings
from scripts.connect_to_database import get_connection
from datetime import datetime, date, timedelta
from openpyxl.styles import PatternFill
from collections import Counter
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    )
from io import BytesIO
import numpy as np


def main(file):
    warnings.filterwarnings('ignore')
    connection = get_connection()
    conexao = pyodbc.connect(connection)

    # Horário
    dt = date.today()
    datetime_midnight = datetime.combine(dt, datetime.min.time())
    date_30 = datetime_midnight - timedelta(30)
    
    # Produtos Materiais
    comando = f'''
    SELECT A.*,  B.COD_INTERNO, B.DESCRICAO, C.ESTOQUE
    FROM ECOM_SKU A
    LEFT JOIN MATERIAIS B ON A.MATERIAL_ID = B.CODID
    LEFT JOIN ESTOQUE_MATERIAIS C ON A.MATERIAL_ID = C.MATERIAL_ID
    WHERE C.ARMAZEM = 1
    AND A.ORIGEM_ID = '8'
    ORDER BY MATERIAL_ID
    '''

    # Carrega planilha
    df_ecom_sku_ml = pd.read_sql(comando, conexao)

    # Limpando valores com espaços vazios
    df_ecom_sku_ml['SKU'] = df_ecom_sku_ml['SKU'].str.strip()
    df_ecom_sku_ml['SKUVARIACAO_MASTER'] = df_ecom_sku_ml['SKUVARIACAO_MASTER'].str.strip()
    df_ecom_sku_ml['DESCRICAO'] = df_ecom_sku_ml['DESCRICAO'].str.strip()

    # Filtra as colunas
    df_ecom_sku_ml = df_ecom_sku_ml[['COD_INTERNO', 'DESCRICAO', 'SKU', 'PRODMKTP_ID', 'SKUVARIACAO_MASTER', 'ORIGEM_ID', 'ESTOQUE']]

    # Renomeia PRODMKTP_ID AS COD_ML
    df_ecom_sku_ml.rename(columns={'PRODMKTP_ID':'COD_ML'}, inplace=True)

    # Preenche os valores vazios da coluna SKU_VARIACAO_MASTER com o valor da coluna SKU
    df_ecom_sku_ml['SKUVARIACAO_MASTER'].fillna(df_ecom_sku_ml['SKU'], inplace=True)
    linhas_vazias = df_ecom_sku_ml.loc[df_ecom_sku_ml['SKUVARIACAO_MASTER'] == '']
    df_ecom_sku_ml.loc[linhas_vazias.index, 'SKUVARIACAO_MASTER'] = linhas_vazias['SKU']

    # Cria a coluna SKU_MESCLADO baseado na coluna SKU que está completa
    df_ecom_sku_ml['SKU_MESCLADO'] = df_ecom_sku_ml['SKUVARIACAO_MASTER']

    # Pedidos
    comando = '''
    SELECT A.PEDIDO, A.COD_INTERNO, A.QUANT, A.COD_PEDIDO AS SKU, A.EDICAO AS SKU2, B.ORIGEM AS ORIGEM_ID, B.DATA
    FROM PEDIDO_MATERIAIS_ITENS_CLIENTE A
    LEFT JOIN PEDIDO_MATERIAIS_CLIENTE B
    ON A.PEDIDO = B.PEDIDO
    WHERE B.TIPO = 'PEDIDO'
    AND B.POSICAO != 'CANCELADO'
    AND B.ORIGEM = '8'
    '''
    data_h = pd.read_sql(comando, conexao)
    data_h.drop('PEDIDO', axis=1, inplace=True)
    data_h['COD_INTERNO'] = data_h['COD_INTERNO'].str.strip()
    data_h['SKU'] = data_h['SKU'].str.strip()
    data_h['SKU2'] = data_h['SKU2'].str.strip()

    # Adiciona pedido
    data_h_aux = data_h[(data_h['QUANT'] > 1)]
    data_h_aux.loc[:, 'QUANT'] -= 1
    for i in range(len(data_h_aux)):
        cod_interno = data_h_aux['COD_INTERNO'].iloc[i]
        quantidade = int(data_h_aux['QUANT'].iloc[i])
        sku = data_h_aux['SKU'].iloc[i]
        sku2 = data_h_aux['SKU2'].iloc[i]
        origem = data_h_aux['ORIGEM_ID'].iloc[i]
        data_venda = data_h_aux['DATA'].iloc[i]
        for j in range(quantidade):
            row1 = pd.Series([cod_interno, 1, sku, sku2, origem, data_venda], index=data_h.columns)
            data_h = data_h._append(row1, ignore_index=True)
            
    # Coloca todos valores da coluna SKU na coluna SKU2 do Mercado Livre
    data_h.loc[data_h['ORIGEM_ID'].isin([8]), 'SKU2'] = data_h['SKU'].fillna(data_h['SKU2'])


    # Fazendo o Groupby de 30 dias
    data_h_30_mktp = data_h[(data_h['DATA'] >= date_30)]
    data_h_30_mktp.loc[:, 'SKU_MESCLADO'] = data_h_30_mktp.apply(lambda x: x['SKU'] if pd.isnull(x['SKU2']) else x['SKU2'], axis=1).copy()
    data_h_30_mktp = data_h_30_mktp.drop(['SKU', 'SKU2', 'DATA'], axis=1)
    data_h_30_mktp = data_h_30_mktp.groupby(['SKU_MESCLADO','ORIGEM_ID']).count()
    data_h_30_mktp = data_h_30_mktp.reset_index()
    data_h_30_mktp = data_h_30_mktp.drop(columns=['COD_INTERNO'], axis=1)

    data_completo = pd.merge(df_ecom_sku_ml, data_h_30_mktp, on=['SKU_MESCLADO', 'ORIGEM_ID'], how='left')
    data_completo['QUANT'].fillna(0, inplace=True)
    data_completo.drop(columns=['SKUVARIACAO_MASTER'], axis=1, inplace=True)
    data_completo = data_completo[['COD_INTERNO', 'DESCRICAO', 'SKU', 'SKU_MESCLADO', 'COD_ML', 'QUANT', 'ESTOQUE']]
    data_completo = data_completo.rename(columns={'QUANT': 'ATON_VENDAS_30', 'SKU': 'MLB_ANUNCIO'})

    # Classificação FULL
    data_completo['E_FULL'] = data_completo['COD_ML'].apply(lambda x: 'FULL' if x != '' and x != None and x != 'null' else 'NAO FULL')

    # Planilha full
    df_ml_full = pd.read_excel(file, skiprows=3, skipfooter=1)
    df_ml_full = df_ml_full.iloc[1:]

    df_ml_full['subtracao'] = df_ml_full['Envios pendentes de recebimento'] + \
                            df_ml_full['Vendas não entregues'] + \
                            df_ml_full['Em transferência'] + \
                            df_ml_full['Devolvidas pelo comprador'] + \
                            df_ml_full['Não aptas para venda'] + \
                            df_ml_full['Aptas para venda']
    df_ml_full['subtracao'] = df_ml_full['subtracao']
    try:
        df_ml_full_filtrada = df_ml_full[['Código ML', 'ID do anúncio', 'Vendas últimos 30 dias (un.)', 'Aptas para venda', 'Tempo até fim do estoque ', 'subtracao']]
        nome_correto = 'Tempo até fim do estoque '
    except:
        df_ml_full_filtrada = df_ml_full[['Código ML', 'ID do anúncio', 'Vendas últimos 30 dias (un.)', 'Aptas para venda', 'Tempo até fim do estoque', 'subtracao']]
        nome_correto = 'Tempo até fim do estoque'

    # Renomeando as colunas
    novos_nomes_colunas = {'Código ML':'COD_ML', 'ID do anúncio':'ID_ANUNCIO', 'Vendas últimos 30 dias (un.)':'VENDAS_30', 'Aptas para venda':'APTAS_FULL', nome_correto:'TEMPO'}
    df_ml_full_filtrada.rename(columns=novos_nomes_colunas, inplace=True)

    # Junta as duas planilhas
    df_completo = pd.merge(data_completo, df_ml_full_filtrada, on='COD_ML', how='left')

    # Cria o campo de ENVIO
    df_completo['ENVIO'] = ''
    df_completo['SUGESTAO'] = ''

    # Altera ordem
    df_completo = df_completo[['COD_INTERNO', 'COD_ML', 'MLB_ANUNCIO', 'SKU_MESCLADO', 'ATON_VENDAS_30', 'VENDAS_30', 'APTAS_FULL', 'ESTOQUE', 'DESCRICAO', 'SUGESTAO', 'ENVIO', 'TEMPO', 'E_FULL', 'subtracao']]
    
    # Preenche os valores vazios com 0
    df_completo['VENDAS_30'].fillna(0, inplace=True)
    df_completo['APTAS_FULL'].fillna(0, inplace=True)

    # Escrever os dataframes em um arquivo Excel com duas abas
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    df_completo.to_excel(writer, sheet_name='ENVIO_FULL', index=False)
    df_ml_full.to_excel(writer, sheet_name='PLAN_FULL', index=False)

    worksheet = writer.sheets['ENVIO_FULL']

    # Adicione o meses
    worksheet['P1'] = 'MESES'
    worksheet['Q1'] = '1'

    # Adiciona a formula nas linhas
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=10, max_col=10): # altera aqui caso adicione uma coluna
        for cell in row:
            cell.value = f'=(E{cell.row}*$Q$1)-N{cell.row}'

    # Alterando tamanho das colunas
    worksheet.column_dimensions['I'].width = '65.43'
    worksheet.column_dimensions['J'].width = '14.86'
    worksheet.column_dimensions['K'].width = '11.43'

    # Adicionando filtros
    worksheet.auto_filter.ref = "A1:N1"

    # # Filtrando valores maior 0 na coluna SUGESTAO
    # filters = worksheet.auto_filter
    # flt1 = CustomFilter(operator="greaterThan", val=0)
    # cfs = CustomFilters(customFilter=[flt1])
    # col = FilterColumn(colId=8, customFilters=cfs)
    # filters.filterColumn.append(col)

    # Escondendo coluna Resultado
    worksheet.column_dimensions['N'].hidden = True

    # # Vendo valores duplicados
    # colunas_alvo = ["B", "H"] 
    # for coluna in colunas_alvo:
    #     valores_coluna = [cell.value for cell in worksheet[coluna]]
    #     valores_duplicados = [valor for valor, frequencia in Counter(valores_coluna).items() if frequencia > 1]
    #     fill = PatternFill(start_color='FFABAB', end_color='FFABAB', fill_type='solid')
    #     for cell in worksheet[coluna]:
    #         if cell.value in valores_duplicados:
    #             cell.fill = fill

    # Congelando painel
    worksheet.freeze_panes = 'A2'
    
    writer._save()
    output.seek(0)

    conexao.close()
    
    return output