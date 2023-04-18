import pyodbc
import pandas as pd
import warnings
from scripts.connect_to_database import get_connection
from openpyxl.styles import PatternFill
from collections import Counter
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    )
from io import BytesIO

def main(file):
    df_ml_full = pd.read_excel(file, skiprows=3, skipfooter=1)
    
    warnings.filterwarnings('ignore')
    connection = get_connection()
    conexao = pyodbc.connect(connection)

    comando = f'''
    SELECT C.DESCRICAO, B.ESTOQUE, A.PRODMKTP_ID AS COD_ML
    FROM ECOM_SKU A
    LEFT JOIN ESTOQUE_MATERIAIS B ON A.MATERIAL_ID = B.MATERIAL_ID
    LEFT JOIN MATERIAIS C ON A.MATERIAL_ID = C.CODID
    WHERE B.ARMAZEM = 1
    AND A.ORIGEM_ID IN('8', '9', '10')
    AND A.PRODMKTP_ID NOT IN('', 'null')
    ORDER BY C.CODID
    '''

    # Materiais
    df_materiais = pd.read_sql(comando, conexao)


    # PLanilha Mercado Livre
    # Lendo o arquivo e pulando as três primeiras linhas
    df_ml_full = df_ml_full.iloc[1:]

    df_ml_full['subtracao'] = df_ml_full['Envios pendentes de recebimento'] + \
                            df_ml_full['Vendas não entregues'] + \
                            df_ml_full['Em transferência'] + \
                            df_ml_full['Devolvidas pelo comprador'] + \
                            df_ml_full['Não aptas para venda'] + \
                            df_ml_full['Aptas para venda']
    df_ml_full['subtracao'] = df_ml_full['subtracao']
    try:
        df_ml_full_filtrada = df_ml_full[['Código ML', 'ID do anúncio', 'Vendas últimos 30 dias (un.)', 'Aptas para venda', 'Tempo até fim do estoque ', 'subtracao']]
        nome_correto = 'Tempo até fim do estoque '
    except:
        df_ml_full_filtrada = df_ml_full[['Código ML', 'ID do anúncio', 'Vendas últimos 30 dias (un.)', 'Aptas para venda', 'Tempo até fim do estoque', 'subtracao']]
        nome_correto = 'Tempo até fim do estoque'

    # Renomeando as colunas
    novos_nomes_colunas = {'Código ML':'COD_ML', 'ID do anúncio':'ID_ANUNCIO', 'Vendas últimos 30 dias (un.)':'VENDAS_30', 'Aptas para venda':'APTAS_FULL', nome_correto:'TEMPO'}
    df_ml_full_filtrada.rename(columns=novos_nomes_colunas, inplace=True)

    # Junta as duas planilhas
    df_completo = pd.merge(df_ml_full_filtrada, df_materiais, on='COD_ML')

    # Cria o campo de ENVIO
    df_completo['ENVIO'] = ''
    df_completo['SUGESTAO'] = ''

    df_completo = df_completo[['COD_ML', 'ID_ANUNCIO', 'VENDAS_30', 'APTAS_FULL', 'ESTOQUE', 'DESCRICAO', 'SUGESTAO', 'ENVIO', 'TEMPO', 'subtracao']]
    
        # Escrever os dataframes em um arquivo Excel com duas abas
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    df_completo.to_excel(writer, sheet_name='ENVIO_FULL', index=False)
    df_ml_full.to_excel(writer, sheet_name='PLAN_FULL', index=False)

    worksheet = writer.sheets['ENVIO_FULL']

    # Adicione o meses
    worksheet['L1'] = 'MESES'
    worksheet['M1'] = '1'

    # Adiciona a formula nas linhas
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=7, max_col=7):
        for cell in row:
            cell.value = f'=(C{cell.row}*$M$1)-J{cell.row}'

    # Alterando tamanho das colunas
    worksheet.column_dimensions['F'].width = '65.43'
    worksheet.column_dimensions['I'].width = '14.86'
    worksheet.column_dimensions['A'].width = '11.43'

    # Adicionando filtros
    worksheet.auto_filter.ref = "A1:J1"

    # Filtrando valores maior 0 na coluna SUGESTAO
    filters = worksheet.auto_filter
    flt1 = CustomFilter(operator="greaterThan", val=0)
    cfs = CustomFilters(customFilter=[flt1])
    col = FilterColumn(colId=6, customFilters=cfs)
    filters.filterColumn.append(col)

    # Resultado
    worksheet.column_dimensions['J'].hidden = True
    
    # Vendo valores duplicados
    colunas_alvo = ["A", "F"] 
    for coluna in colunas_alvo:
        valores_coluna = [cell.value for cell in worksheet[coluna]]
        valores_duplicados = [valor for valor, frequencia in Counter(valores_coluna).items() if frequencia > 1]
        fill = PatternFill(start_color='FFABAB', end_color='FFABAB', fill_type='solid')
        for cell in worksheet[coluna]:
            if cell.value in valores_duplicados:
                cell.fill = fill

    # Congelando painel
    worksheet.freeze_panes = 'A2'
    
    writer._save()
    output.seek(0)
    
    return output