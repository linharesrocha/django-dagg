from scripts.connect_to_database import get_connection
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import pyodbc
from openpyxl.utils import get_column_letter
import pandas as pd
import os
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from dotenv import load_dotenv

def adjust_column_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

def slack_notificao(name_file_excel):
    load_dotenv()

    client = WebClient(token=os.environ['SLACK_BOT_TOKEN'])
    SLACK_CHANNEL_ID='C045HEE4G7L'

    message = f'PLANILHA DE CUSTO TOTAL [ARMAZÉNS]'
    
    # Send message
    try:
        client.chat_postMessage(channel=SLACK_CHANNEL_ID, text=message)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    # Send file
    try:
        client.files_upload_v2(channel=SLACK_CHANNEL_ID, file=name_file_excel, filename=name_file_excel)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))

    # Remove file
    os.remove(name_file_excel)

def main(send_to_slack):
    connection = get_connection()
    conexao = pyodbc.connect(connection)
    
    # Dicionário para mapear números de armazéns para nomes de abas
    armazem_nomes = {
        1: "PRINCIPAL",
        2: "FULL ML MADZ",
        11: "FULL MAGALU MADZ",
        12: "FBA AMAZON",
        13: "FULL MAGALU LEAL"
    }
    
    # Criar um objeto Workbook
    workbook = Workbook()
    
    # Definir a cor laranja para os cabeçalhos
    cor_laranja = PatternFill(start_color='F1C93B', end_color='F1C93B', fill_type='solid')
    
    # Dicionário para armazenar os totais de cada armazém
    resumo_armazens = {}
    
    # Criar as abas de detalhes e coletar dados para o resumo
    for armazem, nome_aba in armazem_nomes.items():
        comando = f'''
        SELECT 
            A.CODID, 
            A.COD_INTERNO, 
            A.DESCRICAO, 
            B.ESTOQUE AS ESTOQUE_REAL, 
            A.VLR_CUSTO,
            A.VLR_CUSTO * B.ESTOQUE AS TOTAL_CUSTO
        FROM MATERIAIS A
        LEFT JOIN ESTOQUE_MATERIAIS B ON A.CODID = B.MATERIAL_ID
        WHERE B.ARMAZEM = {armazem}
        AND A.INATIVO = 'N'
        AND A.COD_INTERNO NOT LIKE '%PAI'
        AND A.COD_INTERNO NOT IN ('7896042027890', '7896042080451')
        AND A.DESMEMBRA = 'N'
        ORDER BY CODID
        '''
        
        df = pd.read_sql(comando, conexao)
        
        # Criar uma nova aba com o nome especificado
        sheet = workbook.create_sheet(title=nome_aba)
        
        # Escrever os cabeçalhos e aplicar a cor laranja
        headers = df.columns.tolist()
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = cor_laranja
        
        # Escrever os dados
        for row, data in enumerate(df.values, start=2):
            for col, value in enumerate(data, start=1):
                sheet.cell(row=row, column=col, value=value)
        
        # Ajustar a largura das colunas
        adjust_column_width(sheet)
        
        # Adicionando filtros
        sheet.auto_filter.ref = sheet.dimensions
        
        # Congelando painel
        sheet.freeze_panes = 'A2'
        
        # Calcular totais para o resumo
        estoque_total = df['ESTOQUE_REAL'].sum()
        valor_total = df['TOTAL_CUSTO'].sum()
        resumo_armazens[nome_aba] = {'ESTOQUE': estoque_total, 'VALOR': valor_total}
    
    # Criar a aba de resumo como primeira aba
    resumo_sheet = workbook.create_sheet(title="RESUMO", index=0)
    
    # Adicionar cabeçalhos ao resumo
    headers_resumo = ["ARMAZEM", "ESTOQUE", "VALOR"]
    for col, header in enumerate(headers_resumo, start=1):
        cell = resumo_sheet.cell(row=1, column=col, value=header)
        cell.fill = cor_laranja
    
    # Preencher dados do resumo
    for row, (armazem, dados) in enumerate(resumo_armazens.items(), start=2):
        resumo_sheet.cell(row=row, column=1, value=armazem)
        resumo_sheet.cell(row=row, column=2, value=dados['ESTOQUE'])
        resumo_sheet.cell(row=row, column=3, value=dados['VALOR'])
    
    # Ajustar a largura das colunas na aba de resumo
    adjust_column_width(resumo_sheet)
    
    # Adicionar filtros à aba de resumo
    resumo_sheet.auto_filter.ref = resumo_sheet.dimensions
    
    # Congelar painel na aba de resumo
    resumo_sheet.freeze_panes = 'A2'
    
    # Remover a planilha padrão criada pelo openpyxl
    workbook.remove(workbook['Sheet'])

    if send_to_slack:
        name_file_excel = 'custo_total_por_armazem.xlsx'
        path_file_excel = os.path.join(os.getcwd(), name_file_excel)
        workbook.save(path_file_excel)

        # Envia slack
        slack_notificao(name_file_excel)
    elif send_to_slack == False:
        return workbook
    
    # Fechar a conexão
    conexao.close()