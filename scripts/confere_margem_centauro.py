import pyodbc
import pandas as pd
import warnings
from datetime import date, timedelta
from dotenv import load_dotenv
from scripts.connect_to_database import get_connection
import os
from slack_sdk import WebClient
import numpy as np
from slack_sdk.errors import SlackApiError
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
import django
from pathlib import Path
import sys
BASE_DIR = Path(__file__).resolve().parent.parent

# Set up Django settings
sys.path.append(str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dagg.settings')
django.setup()

# Import Django models
from trackeamento.models import CentauroTarifas
centauro_tarifas = CentauroTarifas.objects.filter(id=1).first()

def encontra_nome_coluna(sheet, nome_coluna):
    numero_coluna = None
    for coluna in sheet.iter_cols(min_col=1, max_col=sheet.max_column):
        if coluna[0].value == nome_coluna:
            numero_coluna = coluna[0].column
            return numero_coluna

def main(inicio_data_personalizada, fim_data_personalizada, personalizado):
    # Filtrando Warnings
    warnings.filterwarnings('ignore')

    connection = get_connection()
    conexao = pyodbc.connect(connection)
    cursor = conexao.cursor()

    # DAGG
    DAGG_COMISSAO_PADRAO = centauro_tarifas.dagg_comissao_padrao
    DAGG_ACRESCIMO_COMISSAO = 0
    DAGG_DESCONTO_CAMPANHA = 0

    # PADRAO
    TARIFA_FIXA = 0 # PEDIDO ACIMA DE R$10,00
    OPERACAO = 10
    IMPOSTO = 10

    if inicio_data_personalizada == None and fim_data_personalizada == None and personalizado == False:
        comando = f'''
        SELECT DISTINCT A.CODID, C.MATERIAL_ID AS CODID_KIT, A.COD_INTERNO, D.COD_INTERNO AS COD_INTERNO_KIT, COD_PEDIDO AS SKU_MKTP, B.SEUPEDIDO AS PEDIDO, A.EMPRESA, A.DESCRICAOPROD AS TITULO, D.DESCRICAO AS TITULO_KIT, A.VLR_CUSTO, VLR_UNIT AS VLR_PEDIDO, B.VLRFRETE AS VLR_FRETE, DATA, D.DESMEMBRA AS KIT, A.QUANT
        FROM PEDIDO_MATERIAIS_ITENS_CLIENTE A
        LEFT JOIN PEDIDO_MATERIAIS_CLIENTE B ON A.PEDIDO = B.PEDIDO
        LEFT JOIN ECOM_SKU C ON A.COD_PEDIDO = C.SKU
        LEFT JOIN MATERIAIS D ON C.MATERIAL_ID = D.CODID
        WHERE B.TIPO = 'PEDIDO'
        AND B.POSICAO != 'CANCELADO'
        AND B.ORIGEM IN ('34')
        AND B.DATA >= DATEADD(DAY, DATEDIFF(DAY, 0, GETDATE()) - 6, 0)
        AND B.DATA < DATEADD(DAY, DATEDIFF(DAY, 0, GETDATE()), 0)
        ORDER BY A.EMPRESA, DATA
        '''
    else:
        comando = f'''
        SELECT DISTINCT A.CODID, C.MATERIAL_ID AS CODID_KIT, A.COD_INTERNO, D.COD_INTERNO AS COD_INTERNO_KIT, COD_PEDIDO AS SKU_MKTP, B.SEUPEDIDO AS PEDIDO, A.EMPRESA, A.DESCRICAOPROD AS TITULO, D.DESCRICAO AS TITULO_KIT, A.VLR_CUSTO, VLR_UNIT AS VLR_PEDIDO, B.VLRFRETE AS VLR_FRETE, DATA, D.DESMEMBRA AS KIT, A.QUANT
        FROM PEDIDO_MATERIAIS_ITENS_CLIENTE A
        LEFT JOIN PEDIDO_MATERIAIS_CLIENTE B ON A.PEDIDO = B.PEDIDO
        LEFT JOIN ECOM_SKU C ON A.COD_PEDIDO = C.SKU
        LEFT JOIN MATERIAIS D ON C.MATERIAL_ID = D.CODID
        WHERE B.TIPO = 'PEDIDO'
        AND B.POSICAO != 'CANCELADO'
        AND B.ORIGEM IN ('34')
        AND CONVERT(DATE, B.DATA) >= '{inicio_data_personalizada}'
        AND CONVERT(DATE, B.DATA) <= '{fim_data_personalizada}'
        ORDER BY A.EMPRESA, DATA
        '''

    df = pd.read_sql(comando, conexao)

    # Remove espaço da coluna TITULO
    df['TITULO'] = df['TITULO'].str.strip()
    df['TITULO_KIT'] = df['TITULO_KIT'].str.strip()
    df['COD_INTERNO'] = df['COD_INTERNO'].str.strip()
    df['COD_INTERNO_KIT'] = df['COD_INTERNO_KIT'].str.strip()
    df['PEDIDO'] = df['PEDIDO'].str.strip()

    is_kit_s = df['KIT'] == 'S'

    # Altera CODID_KIT
    df['CODID'] = np.where(is_kit_s, df['CODID_KIT'], df['CODID'])
    df.drop('CODID_KIT', axis=1, inplace=True)

    # Altera COD_INTERNO_KIT
    df['COD_INTERNO'] = np.where(is_kit_s, df['COD_INTERNO_KIT'], df['COD_INTERNO'])
    df.drop('COD_INTERNO_KIT', axis=1, inplace=True)

    # Altera TITULO_KIT
    df['TITULO'] = np.where(is_kit_s, df['TITULO_KIT'], df['TITULO'])
    df.drop('TITULO_KIT', axis=1, inplace=True)

    # Somando VLR_CUSTO e VLR_PEDIDO
    df['SOMA_VLR_CUSTO'] = 0
    df['SOMA_VLR_PEDIDO'] = 0
    mask_custo = (df['KIT'] == 'S')
    mask_pedido = (df['KIT'] == 'S')
    soma_custo = df[mask_custo].groupby('PEDIDO')['VLR_CUSTO'].transform(lambda x: (x * df.loc[x.index, 'QUANT']).sum())
    soma_pedido = df[mask_pedido].groupby('PEDIDO')['VLR_PEDIDO'].transform(lambda x: (x * df.loc[x.index, 'QUANT']).sum())
    df.loc[mask_custo, 'SOMA_VLR_CUSTO'] = soma_custo
    df.loc[mask_pedido, 'SOMA_VLR_PEDIDO'] = soma_pedido
    df['VLR_CUSTO'] = df['SOMA_VLR_CUSTO'].where(mask_custo, df['VLR_CUSTO'])
    df['VLR_PEDIDO'] = df['SOMA_VLR_PEDIDO'].where(mask_pedido, df['VLR_PEDIDO'])
    df.drop(columns=['SOMA_VLR_CUSTO', 'SOMA_VLR_PEDIDO'], inplace=True)

    # Removendo duplicadas
    df_s = df[df['KIT'] == 'S']
    df_s_no_duplicates = df_s.drop_duplicates()
    df = pd.concat([df[df['KIT'] != 'S'], df_s_no_duplicates])

    # Quantidade Pedidos
    df['QUANTIDADE_PED'] = df.groupby(['EMPRESA', 'PEDIDO'])['PEDIDO'].transform('count')

    # Arrendonda coluna VLR_PEDIDO para 2 casas decimais
    df['VLR_PEDIDO'] = round(df['VLR_PEDIDO'], 2)

    # Adiciona coluna VLR_TOTAL
    df['VLR_TOTAL'] = df['VLR_PEDIDO'] + df['VLR_FRETE']

    # Coloca a coluna VLR_TOTAL depois do VLR_FRETE
    df = df[['CODID', 'COD_INTERNO', 'PEDIDO', 'EMPRESA', 'TITULO', 'VLR_CUSTO', 'VLR_PEDIDO', 'VLR_FRETE', 'VLR_TOTAL', 'DATA', 'KIT','QUANTIDADE_PED']]

    # Modifica VLR_FRETE
    df['VLR_FRETE'] = (df['VLR_FRETE'] / df['QUANTIDADE_PED']).round(2)

    # Adiciona coluna TARIFA FIXA para apenas os valores da coluna VLR_PEDIDO acima de R$10,00 se não 0
    df['TARIFA_FIXA'] = df['VLR_PEDIDO'].apply(lambda x: TARIFA_FIXA if x > 10 else 0)

    # Modifica TARIFICA_FIXA
    df['TARIFA_FIXA'] = (df['TARIFA_FIXA'] / df['QUANTIDADE_PED']).round(2)

    # Imposto Frete que sigifnica aba IMPOSSTO * VALOR DO FRETE
    df['IMPOSTO_FRETE'] = (df['VLR_FRETE'] * (IMPOSTO / 100)).round(2)

    # Adiciona coluna OPERACAO
    df['OPERACAO'] = centauro_tarifas.imposto

    # Adiciona coluna IMPOSTO
    df['IMPOSTO'] = centauro_tarifas.imposto

    # Adiciona coluna COMISSAO_PADRAO mas diferenciando por empresa
    df['COMISSAO_PADRAO'] = DAGG_COMISSAO_PADRAO

    # Adiciona coluna ACRESCIMO_COMISSAO mas diferenciando por empresa
    df['ACRESCIMO_COMISSAO'] = DAGG_ACRESCIMO_COMISSAO

    # Adiciona coluna DESCONTO_CAMPANHA mas diferenciando por empresa
    df['DESCONTO_CAMPANHA'] = DAGG_DESCONTO_CAMPANHA

    # Adiciona coluna PORC_TOTAL que soma as colunas
    df['CUSTO_PARCIAL%'] = df['OPERACAO'] + df['IMPOSTO'] + df['COMISSAO_PADRAO'] + df['ACRESCIMO_COMISSAO'] - df['DESCONTO_CAMPANHA']

    # Calculo resto
    df['CUSTO_PARCIAL$'] = round(df['VLR_PEDIDO'] * (df['CUSTO_PARCIAL%'] / 100), 2)

    # Lucro
    df['LUCRO'] = round(df['VLR_PEDIDO'] - df['CUSTO_PARCIAL$'] - df['TARIFA_FIXA'] - df['VLR_CUSTO'] - df['IMPOSTO_FRETE'], 2)

    # Margem
    df['MARGEM'] = round(df['LUCRO'] / df['VLR_PEDIDO'] * 100, 2)

    # Altera coluna EMPRESA para nomes
    df['EMPRESA'] = df['EMPRESA'].replace(1, 'DAGG')
    df['EMPRESA'] = df['EMPRESA'].replace(2, 'RED PLACE')
    df['EMPRESA'] = df['EMPRESA'].replace(3, 'PISSTE')

    # Salve em excel com a data de ontem
    if personalizado == False:
        name_file_excel = 'margem_centauro_' + str(date.today() - timedelta(days=1)) + '.xlsx'
    else:
        name_file_excel = 'margem_centauro_personalizado'+ '.xlsx'

    writer = pd.ExcelWriter(name_file_excel, engine='openpyxl')
    df.to_excel(writer, sheet_name='centauro', index=False)

    worksheet = writer.sheets['centauro']

    # Adicionando filtros
    worksheet.auto_filter.ref = "A1:W1"

    # Congelando painel
    worksheet.freeze_panes = 'A2'

    # Definir a cor
    cor_laranja = PatternFill(start_color='F1C93B', end_color='F1C93B', fill_type='solid')
    cor_verde = PatternFill(start_color='96C291', end_color='96C291', fill_type='solid')
    cor_branco = PatternFill(start_color='F4EEEE', end_color='F4EEEE', fill_type='solid')

    # Lista Porcentagem laranja
    celulas_laranja = ['W1', 'T1', 'S1', 'R1', 'Q1', 'P1', 'O1']
    for celula_referencia in celulas_laranja:
        celula = worksheet[celula_referencia]
        celula.fill = cor_laranja
        
    # Lista Porcentagem Verde
    celulas_verde = ['V1', 'U1', 'N1', 'M1', 'I1', 'H1', 'G1', 'F1']
    for celula_referencia in celulas_verde:
        celula = worksheet[celula_referencia]
        celula.fill = cor_verde
        

    # Lista Porcentagem Branco
    celulas_branco = ['K1', 'L1', 'E1', 'D1', 'C1', 'B1', 'A1', 'J1']
    for celula_referencia in celulas_branco:
        celula = worksheet[celula_referencia]
        celula.fill = cor_branco


    writer._save()

    # Envia slack
    load_dotenv()

    client = WebClient(token=os.environ['SLACK_BOT_TOKEN'])
    SLACK_CHANNEL_ID='C07HKDM1PFD'

    message = f'centauro MARGEM! :heavy_division_sign:'

    # Send message
    try:
        client.chat_postMessage(channel=SLACK_CHANNEL_ID, text=message)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    # Send file
    try:
        client.files_upload_v2(channel=SLACK_CHANNEL_ID, file=name_file_excel, filename=name_file_excel)
    except SlackApiError as e:
        print("Error sending message: {}".format(e))
        
    writer.close()
        
    # Remove arquivo
    try:
        os.remove(name_file_excel)
    except Exception as e:
        print(e)